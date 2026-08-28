"""Exports the master workbook's data sheets to `data/*.csv`.

    python tools/export_workbook.py [--check]

Authorization §15: once the director imports the text sheets, the workbook is
the authoring source and the build consumes exported CSVs. This makes that step
reproducible instead of manual — a hand export is where a column quietly goes
missing, which is exactly what happened on the first pass (BOS lost its primary
key between the workbook and the eye that checked it).

`--check` reports drift without writing, so a stale `data/` can be detected in a
verification run rather than discovered at load.

Non-data sheets (notes, conventions) are ignored: they are authoring aids and
have never shipped.
"""

import argparse
import csv
import io
import pathlib
import sys

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = pathlib.Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "staging" / "breach datastructures.xlsx"
DATA = ROOT / "data"

# Sheet -> filename. Only these ship; anything else in the workbook is notes.
SHEETS = {
    "PRG_H": "prg_h.csv",
    "PRG_S": "prg_s.csv",
    "FNC": "fnc.csv",
    "HAK": "hak.csv",
    "PSV": "psv.csv",
    "DEK": "dek.csv",
    "SYS": "sys.csv",
    "HST": "hst.csv",
    "UPG": "upg.csv",
    "BOS": "bos.csv",
    "text_content": "text_content.csv",
    "text_style": "text_style.csv",
    "font_refs": "font_refs.csv",
}


def cell(v):
    """Excel gives numbers back as floats; the CSVs are authored as integers.

    `250.0` in a BASE_ICE column would fail the loader's integer parse, so a
    whole-number float is written back as an integer. A genuinely fractional
    value is left alone rather than silently truncated.
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    # Trailing line breaks only. NOT .strip(): whitespace inside a text_content
    # cell is content, and stripping it here would compound a loss rather than
    # cause one — Excel already discards LEADING spaces on CSV import, which is
    # how the battle log's indented sub-messages were flattened. See AN-011.
    return str(v).rstrip(chr(13) + chr(10))


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [cell(h) for h in rows[0]]
    while header and header[-1] == "":
        header.pop()

    body = []
    for r in rows[1:]:
        vals = [cell(c) for c in r][: len(header)]
        while len(vals) < len(header):
            vals.append("")
        if any(v for v in vals):
            body.append(vals)
    return header, body


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="report drift, write nothing")
    args = ap.parse_args()

    if not WORKBOOK.exists():
        print("no workbook at %s" % WORKBOOK)
        return 1

    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    drift = 0

    for sheet, fname in SHEETS.items():
        if sheet not in wb.sheetnames:
            print("  MISSING SHEET %s" % sheet)
            drift += 1
            continue

        header, body = sheet_rows(wb[sheet])
        if not header:
            print("  EMPTY SHEET %s" % sheet)
            drift += 1
            continue

        buf = io.StringIO()
        w = csv.writer(buf, lineterminator="\n")
        w.writerow(header)
        w.writerows(body)
        text = buf.getvalue()

        path = DATA / fname
        current = path.read_text(encoding="utf-8") if path.exists() else ""

        if current == text:
            print("  same  %-20s %d rows" % (fname, len(body)))
            continue

        drift += 1
        if args.check:
            print("  DRIFT %-20s %d rows" % (fname, len(body)))
        else:
            path.write_text(text, encoding="utf-8", newline="")
            print("  wrote %-20s %d rows" % (fname, len(body)))

    if args.check and drift:
        print("\n%d sheet(s) differ from data/ — run without --check to export" % drift)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
